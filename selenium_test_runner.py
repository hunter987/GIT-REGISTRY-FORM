import os
import threading
import time
import re
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter import Tk, filedialog
import validators
# ---- Funciones de validación personalizadas (ejemplo básico) ----
def validar_nombre(name):
    return bool(name.strip())

def validar_email(email):
    return "@" in email and "." in email

def validar_password(password):
    # Ejemplo: mínimo 6 caracteres, al menos una mayúscula, una minúscula y un número
    import re
    if len(password) < 6:
        return False
    if not re.search(r"[A-Z]", password):
        return False
    if not re.search(r"[a-z]", password):
        return False
    if not re.search(r"\d", password):
        return False
    return True

def validar_confirmacion(password, confirm):
    return password == confirm

def evaluar_validadores(name, email, password, confirm):
    return {
        "Valida Nombre": "OK" if validar_nombre(name) else "Falla",
        "Valida Email": "OK" if validar_email(email) else "Falla",
        "Valida Password": "OK" if validar_password(password) else "Falla",
        "Confirma Password": "OK" if validar_confirmacion(password, confirm) else "Falla",
    }

# --- Función para seleccionar archivo Excel ---
def seleccionar_archivo():
    root = Tk()
    root.withdraw()
    archivo = filedialog.askopenfilename(
        title="Selecciona archivo Excel con datos de usuarios",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")]
    )
    root.destroy()
    return archivo

# --- Función para iniciar servidor Flask (simulado aquí) ---
def run_flask_app():
    # Aquí va el código de tu servidor Flask que sirva el form.html y reciba datos.
    # Se asume que ya está corriendo y accesible en http://127.0.0.1:5000/form.html
    pass

# --- Función principal de ejecución de pruebas con Selenium ---
def ejecutar_registros():
    archivo = seleccionar_archivo()
    if not archivo:
        print("No se seleccionó archivo.")
        return

    # Si usas un servidor Flask para servir la página, puedes iniciarlo en otro hilo.
    flask_thread = threading.Thread(target=run_flask_app, daemon=True)
    flask_thread.start()
    print("Servidor Flask arrancado en hilo separado...")
    time.sleep(3)

    wb = load_workbook(archivo)
    ws = wb.active
    encabezados = ["Full Name", "Email", "Password", "Confirm Password", "Expected Outcome", "Resultado del Test", "Coincidencia con Esperado"]
    for idx, titulo in enumerate(encabezados, start=1):
        if not ws.cell(row=1, column=idx).value:
            ws.cell(row=1, column=idx).value = titulo

    driver_path = os.path.abspath(os.path.join("chromedriver-win64", "chromedriver.exe"))
    service = Service(driver_path)
    options = webdriver.ChromeOptions()
    options.add_argument("--log-level=3")
    driver = webdriver.Chrome(service=service, options=options)

    url_form = "http://127.0.0.1:5000/form.html"
    wait = WebDriverWait(driver, 5)

    print(f"Iniciando procesamiento de {ws.max_row - 1} registros...")

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 4:
            print(f"Fila {i} incompleta. Saltando...")
            continue

        name, email, password, confirm = row[:4]
        expected = row[4] if len(row) >= 5 else None

        resultado = "No Ejecutado"

        try:
            driver.get(url_form)

            fullname_input = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@id='fullname' or @name='fullname']")))
            fullname_input.clear()
            fullname_input.send_keys(name)

            email_input = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@id='email' or @name='email']")))
            email_input.clear()
            email_input.send_keys(email)

            password_input = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@id='password' or @name='password']")))
            password_input.clear()
            password_input.send_keys(password)

            confirm_input = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@id='confirm_password' or @name='confirm_password']")))
            confirm_input.clear()
            confirm_input.send_keys(confirm)

            submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='submit' or @type='submit']")))
            submit_button.click()

            validation_message_elem = wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@id='validation_message' or contains(@class, 'validation_message')]")))
            validation_message = validation_message_elem.text.strip()

            resultado = validation_message

        except Exception as e:
            resultado = f"Error Selenium: {str(e)}"

        ws.cell(row=i, column=6).value = resultado

        if expected:
            match = "✅" if expected.strip().lower() in resultado.lower() else "❌"
            ws.cell(row=i, column=7).value = match

    driver.quit()
    salida = os.path.join(os.path.dirname(archivo), "resultados.xlsx")
    wb.save(salida)
    print(f"✅ Finalizado. Resultados guardados en {salida}.")

    # Genera archivo detallado con resultados individuales
    detalle_path = os.path.join(os.path.dirname(archivo), "resultados_detalle_validadores.xlsx")
    procesar_resultados_excel(salida, detalle_path)
    print(f"✅ Archivo de detalle generado en {detalle_path}.")

# --- Función para crear el Excel con resultados de validadores individuales ---
def procesar_resultados_excel(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active

    wb_out = Workbook()
    ws_out = wb_out.active

    encabezados = ["Full Name", "Email", "Password", "Confirm Password", "Expected Outcome"]
    validador_headers = list(validators.evaluar_validadores("Test", "test@example.com", "Pass1A$", "Pass1A$").keys())
    ws_out.append(encabezados + validador_headers)

    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre, correo, clave, confirmacion = row[:4]
        resultado_validadores = validators.evaluar_validadores(nombre or "", correo or "", clave or "", confirmacion or "")
        fila = [
            nombre, correo, clave, confirmacion,
            row[4] if len(row) >= 5 else ""
        ] + list(resultado_validadores.values())
        ws_out.append(fila)

    wb_out.save(output_path)

# --- Código principal ---
if __name__ == "__main__":
    ejecutar_registros()
